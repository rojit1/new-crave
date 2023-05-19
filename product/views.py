from django.shortcuts import redirect, render, get_object_or_404
from openpyxl import load_workbook
from organization.models import Branch

from django.urls import reverse_lazy
from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    UpdateView,
    View,
)
from root.utils import DeleteMixin
from user.permission import IsAdminMixin
from .models import ProductCategory, ItemReconcilationApiItem
from bill.models import Bill
from .forms import ProductCategoryForm
import datetime


class ProductCategoryMixin(IsAdminMixin):
    model = ProductCategory
    form_class = ProductCategoryForm
    paginate_by = 50
    queryset = ProductCategory.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("product_category_list")
    search_lookup_fields = [
        "title",
        "description",
    ]


class ProductCategoryList(ProductCategoryMixin, ListView):
    template_name = "productcategory/productcategory_list.html"
    queryset = ProductCategory.objects.filter(status=True, is_deleted=False)


class ProductCategoryDetail(ProductCategoryMixin, DetailView):
    template_name = "productcategory/productcategory_detail.html"


class ProductCategoryCreate(ProductCategoryMixin, CreateView):
    template_name = "create.html"


class ProductCategoryUpdate(ProductCategoryMixin, UpdateView):
    template_name = "update.html"


class ProductCategoryDelete(ProductCategoryMixin, DeleteMixin, View):
    pass


from django.db import transaction
from django.http import JsonResponse
from django.urls import reverse_lazy
from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    TemplateView,
    UpdateView,
    View,
)
from root.utils import DeleteMixin
from .models import Product
from .forms import ProductForm


class ProductMixin(IsAdminMixin):
    model = Product
    form_class = ProductForm
    paginate_by = 50
    queryset = Product.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("product_list")
    search_lookup_fields = [
        "title",
        "description",
    ]


class ProductList(ProductMixin, ListView):
    template_name = "product/product_list.html"
    queryset = Product.objects.filter(status=True, is_deleted=False)


class ProductDetail(ProductMixin, DetailView):
    template_name = "product/product_detail.html"


class ProductCreate(ProductMixin, CreateView):
    template_name = "product/product_create.html"


class ProductUpdate(ProductMixin, UpdateView):
    template_name = "update.html"


class ProductDelete(ProductMixin, DeleteMixin, View):
    pass

class ProductUploadView(View):

    def post(self, request):
        file = request.FILES['file']
        wb = load_workbook(file)
        food_category, _ = ProductCategory.objects.get_or_create(title='FOOD')
        beverage_category, _ = ProductCategory.objects.get_or_create(title='BEVERAGE')
        others_category, _ = ProductCategory.objects.get_or_create(title='OTHERS')
        excel_data = list()
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                row_data.insert(0, sheet.title)
                excel_data.append(row_data)
        for data in excel_data:
            if data[1].lower() == 'group':
                continue
            try:
                product = Product.objects.get(title__iexact=data[2])
                product.group = data[1]
                product.price = data[3]
                product.unit = data[4]
                product.is_taxable = True if data[5].lower() == "yes" else False
                product.is_produced = True if data[6].lower() == "yes" else False
                product.reconcile = True if data[7].lower() == "yes" else False
                if "food" in data[0].lower():
                    product.type = food_category
                elif "beverage" in data[0].lower():
                    product.type = beverage_category
                else:
                    product.type = others_category
                product.save()
            except Product.DoesNotExist:
                pass
                product = Product()
                product.group = data[1]
                product.title=data[2]
                product.price = data[3]
                product.unit = data[4]
                product.is_taxable = True if data[5].lower() == "yes" else False
                product.is_produced = True if data[6].lower() == "yes" else False
                product.reconcile = True if data[7].lower() == "yes" else False
                if "food" in data[0].lower():
                    product.type = food_category
                elif "beverage" in data[0].lower():
                    product.type = beverage_category
                else:
                    product.type = others_category
                product.save()

        return redirect(reverse_lazy("product_list"))

from django.urls import reverse_lazy
from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    TemplateView,
    UpdateView,
    View,
)
from root.utils import DeleteMixin
from .models import CustomerProduct
from .forms import CustomerProductForm


class CustomerProductMixin(IsAdminMixin):
    model = CustomerProduct
    form_class = CustomerProductForm
    paginate_by = 50
    queryset = CustomerProduct.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("customerproduct_list")
    search_lookup_fields = ["product__title", "customer__name", "agent__full_name"]


class CustomerProductList(CustomerProductMixin, ListView):
    template_name = "customerproduct/customerproduct_list.html"
    queryset = CustomerProduct.objects.filter(status=True, is_deleted=False)


class CustomerProductDetail(CustomerProductMixin, DetailView):
    template_name = "customerproduct/customerproduct_detail.html"


class CustomerProductCreate(CustomerProductMixin, CreateView):
    template_name = "create.html"

    def form_valid(self, form):
        form.instance.agent = self.request.user
        return super().form_valid(form)


class CustomerProductUpdate(CustomerProductMixin, UpdateView):
    template_name = "update.html"

    def form_valid(self, form):
        form.instance.agent = self.request.user
        return super().form_valid(form)


class CustomerProductDelete(CustomerProductMixin, DeleteMixin, View):
    pass

'''  STock VIews '''

from .models import ProductStock
from .forms import ProductStockForm

class ProductStockMixin:
    model = ProductStock
    form_class = ProductStockForm
    paginate_by = 10
    queryset = ProductStock.objects.filter(status=True,is_deleted=False)
    success_url = reverse_lazy('productstock_list')


class ProductStockList(ProductStockMixin, ListView):
    template_name = "productstock/productstock_list.html"
    queryset = ProductStock.objects.filter(status=True,is_deleted=False)

class ProductStockDetail(ProductStockMixin, DetailView):
    template_name = "productstock/productstock_detail.html"

class ProductStockCreate(ProductStockMixin, CreateView):
    template_name = "create.html"

class ProductStockUpdate(ProductStockMixin, UpdateView):
    template_name = "update.html"

class ProductStockDelete(ProductStockMixin, DeleteMixin, View):
    pass

"""  ----------------   """
from .models import BranchStock, BranchStockTracking
from .forms import BranchStockForm
class BranchStockMixin:
    model = BranchStock
    form_class = BranchStockForm
    paginate_by = 10
    queryset = BranchStock.objects.filter(status=True,is_deleted=False)
    success_url = reverse_lazy('branchstock_list')

class BranchStockList(BranchStockMixin, ListView):
    template_name = "branchstock/branchstock_list.html"

class BranchStockDetail(BranchStockMixin, DetailView):
    template_name = "branchstock/branchstock_detail.html"

class BranchStockCreate(BranchStockMixin, CreateView):
    template_name = "branchstock/branchstock_create.html"

class BranchStockUpdate(BranchStockMixin, UpdateView):
    template_name = "update.html"

class BranchStockDelete(BranchStockMixin, DeleteMixin, View):
    pass


from datetime import date, datetime
from django.db.models import Sum
class ReconcileView(View): 

    def get(self, request):

        branch = request.GET.get('branch', None)
        filter_date = request.GET.get('date')
        branches = Branch.objects.all()
        if not branch:
            return render(request, 'item_reconcilation/reconcilation.html',{'message':'Please Select a Branch', 'branches':branches})
        if not filter_date:
            filter_date = date.today().strftime('%Y-%m-%d')
        else:
            filter_date = datetime.strptime(filter_date, '%Y-%m-%d').date().strftime('%Y-%m-%d')
        filter_branch = get_object_or_404(Branch, branch_code__iexact=branch)
        


        if filter_date == date.today().strftime('%Y-%m-%d'):
            products = Product.objects.filter(reconcile=True).order_by('title').values()
            api_items = ItemReconcilationApiItem.objects.filter(date=filter_date, branch=filter_branch, product__reconcile=True).values()
            received = BranchStock.objects.filter(created_at__contains=filter_date, branch=filter_branch, product__reconcile=True).values('product').annotate(quantity=Sum('quantity'))
            bills = Bill.objects.filter(transaction_date=filter_date, branch=filter_branch)

            new_products = {}
            for product in products:
                for k, v in product.items():
                    if k =='id':
                        new_products[str(v)] = {'title':product.get('title')}
                        break
       
            for item in api_items:
                product_id = str(item.get('product_id'))
                new_products[product_id]['wastage'] = item.get('wastage', 0)
                new_products[product_id]['returned'] = item.get('returned', 0)
                new_products[product_id]['physical'] = item.get('physical', 0)

            for rec in received:
                product_id = str(rec.get('product'))
                new_products[product_id]['received'] = rec.get('quantity')
            
            for bill in bills:
                for item in bill.bill_items.all():
                    product_id = str(item.product_id)
                    if item.product.reconcile:
                        has_sold = new_products[product_id].get('sold', None)
                        if has_sold:
                            new_products[product_id]['sold'] += item.product_quantity
                        else:
                            new_products[product_id]['sold'] = item.product_quantity

            product_to_view = []
            for k,v in new_products.items():
                new_dict = {'id': k, **v}
                if not 'opening' in new_dict:
                    new_dict['opening'] = 0
                if not 'received' in new_dict:
                    new_dict['received'] = 0
                if not 'wastage' in new_dict:
                    new_dict['wastage'] = 0
                if not 'returned' in new_dict:
                    new_dict['returned'] = 0
                if not 'sold' in new_dict:
                    new_dict['sold'] = 0
                if not 'closing' in new_dict:
                    new_dict['closing'] = 0
                if not 'physical' in new_dict:
                    new_dict['physical'] = 0
                if not 'discrepancy' in new_dict:
                    new_dict['discrepancy'] = 0

                product_to_view.append(new_dict)
            
            for prd in product_to_view:
                opening_received = prd.get('opening') + prd.get('received')
                wastage_returned_sold = prd.get('wastage') + prd.get('returned') + prd.get('sold')
                closing_value = opening_received - wastage_returned_sold
                prd['closing'] = closing_value
                prd['discrepancy'] = prd.get('physical') - closing_value

            context = {
                'products':product_to_view,
                'branches':branches,
                'should_save':True
            }
            return render(request, 'item_reconcilation/reconcilation.html',context)
        # --------------------------


        products = BranchStockTracking.objects.filter(date=filter_date)
        context = {
            'products':products,
            'branches':branches,
            'should_save':False
        }
        return render(request, 'item_reconcilation/reconcilation.html', context)
    
    def post(self, request):
        print(request.POST)
        return render(request, 'item_reconcilation/reconcilation.html', {'branches':Branch.objects.all()})


from django.contrib import messages
class BranchStockUploadView(View):

    def post(self, request):
        return redirect(reverse_lazy("branchstock_create"))
        if BranchStockTracking.objects.first():
            messages.error(request, "Opening data already exists!!")
            return redirect(reverse_lazy("branchstock_create"))
        file = request.FILES.get('file')
        branches = Branch.objects.all()
        branch_dict = {}
        for b in branches:
            branch_dict[b.branch_code.lower()] = b.pk

        wb = load_workbook(file)
        excel_data = list()
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)
        excel_data.pop(0)

        product_dict = {}

        for d in excel_data:
            try:
                product_title = d[2].lower()
                product = product_dict.get(product_title, None)
                if not product:
                    product_dict[product_title] = Product.objects.get(title__iexact=product_title).pk
                product_id = product_dict.get(product_title)
                branch_id =  branch_dict.get(d[1].lower())
                quantity = int(d[3])
                formatted_date = d[0]
            except Exception as e:
                print(e)

        return redirect(reverse_lazy("branchstock_create"))